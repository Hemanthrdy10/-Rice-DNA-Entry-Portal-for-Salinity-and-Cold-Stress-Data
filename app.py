from flask import Flask, render_template, request, redirect, url_for, session, send_file
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)
app.secret_key = 'secret_key_123'  # Required for session management

# Dummy login credentials
USER_CREDENTIALS = {'pavan': 'pavan@123'}

@app.route('/')
def home():
    if 'user' not in session:
        return redirect(url_for('login'))
    return render_template('index.html', message=None)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        if username in USER_CREDENTIALS and USER_CREDENTIALS[username] == password:
            session['user'] = username
            session['excel_file'] = 'data.xlsx'  # Default file on login

            # Create the default file if it doesn't exist
            if not os.path.exists('data.xlsx'):
                wb = Workbook()
                ws = wb.active
                ws.append(["Name", "Feature2", "Feature3", "Feature4", "Feature5"])
                wb.save('data.xlsx')

            return redirect(url_for('home'))
        else:
            return render_template('login.html', error="Invalid credentials")
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('user', None)
    session.pop('excel_file', None)
    return redirect(url_for('login'))

@app.route('/submit', methods=['POST'])
def submit():
    if 'user' not in session:
        return redirect(url_for('login'))

    filename = session.get('excel_file', 'data.xlsx')
    f1 = request.form['Name']
    f2 = request.form['feature2']
    f3 = request.form['feature3']
    f4 = request.form['feature4']
    f5 = request.form['feature5']

    wb = load_workbook(filename)
    ws = wb.active
    ws.append([f1, f2, f3, f4, f5])
    wb.save(filename)

    return render_template('index.html', message="✅ Data saved!")

@app.route('/download')
def download():
    if 'user' not in session:
        return redirect(url_for('login'))
    filename = session.get('excel_file', 'data.xlsx')
    return send_file(filename, as_attachment=True)

@app.route('/create_excel', methods=['POST'])
def create_excel():
    if 'user' not in session:
        return redirect(url_for('login'))

    filename = request.form['filename']
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'

    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Feature2", "Feature3", "Feature4", "Feature5"])
        wb.save(filename)
        message = f"✅ Excel file '{filename}' created and switched!"
    else:
        message = f"🔁 Switched to existing file '{filename}'."

    session['excel_file'] = filename  # Switch file
    return render_template('index.html', message=message)

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')
